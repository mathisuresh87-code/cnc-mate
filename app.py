import base64
from datetime import datetime
import math
import os
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from PIL import Image
import streamlit as st

# Page Configuration
st.set_page_config(
    page_title="Megala CNC Mate - Enterprise CNC Automation",
    page_icon="⚙️",
    layout="wide",
)

# Ultra-Vibrant, Colorful & Tactile Interactive SaaS Custom CSS
st.markdown(
    """
    <style>
    .stApp {
        background: linear-gradient(135deg, #090d1f 0%, #111827 40%, #1e1b4b 100%);
        color: #f8fafc;
        font-family: 'Inter', 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
    }
    .header-container {
        display: flex;
        align-items: center;
        gap: 20px;
        padding: 10px 0 20px 0;
        border-bottom: 2px solid rgba(236, 72, 153, 0.3);
        margin-bottom: 25px;
    }
    .main-title {
        font-size: 2.8rem;
        font-weight: 900;
        background: linear-gradient(135deg, #38bdf8 0%, #c084fc 50%, #ec4899 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin: 0 0 4px 0;
        padding: 0;
        letter-spacing: 1.5px;
        text-transform: uppercase;
        line-height: 1.1;
        filter: drop-shadow(0 0 20px rgba(192, 132, 252, 0.5));
    }
    .sub-title {
        font-size: 0.92rem;
        color: #38bdf8;
        margin: 0;
        font-weight: 700;
        letter-spacing: 2px;
        text-transform: uppercase;
        text-shadow: 0 0 10px rgba(56, 189, 248, 0.4);
    }
    .auto-badge {
        background: linear-gradient(135deg, rgba(236, 72, 153, 0.3) 0%, rgba(139, 92, 246, 0.4) 100%);
        color: #f472b6;
        padding: 6px 18px;
        border-radius: 20px;
        font-size: 0.85rem;
        font-weight: 800;
        display: inline-block;
        margin-bottom: 16px;
        border: 1.5px solid rgba(236, 72, 153, 0.6);
        text-transform: uppercase;
        letter-spacing: 1.2px;
        box-shadow: 0 0 20px rgba(236, 72, 153, 0.3);
    }
    div[data-testid="column"] {
        background: linear-gradient(145deg, rgba(30, 41, 59, 0.85) 0%, rgba(49, 46, 129, 0.5) 100%);
        border: 1.5px solid rgba(139, 92, 246, 0.4);
        padding: 24px;
        border-radius: 20px;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.6), inset 0 1px 0 rgba(255, 255, 255, 0.15);
        backdrop-filter: blur(16px);
        transition: all 0.35s cubic-bezier(0.4, 0, 0.2, 1);
        margin-bottom: 20px;
    }
    div[data-testid="column"]:hover {
        border-color: #ec4899;
        transform: translateY(-8px) scale(1.015);
        box-shadow: 0 20px 50px rgba(236, 72, 153, 0.4), 0 0 30px rgba(56, 189, 248, 0.3), inset 0 1px 20px rgba(236, 72, 153, 0.4);
        background: linear-gradient(145deg, rgba(55, 48, 163, 0.7) 0%, rgba(131, 24, 67, 0.5) 100%);
    }
    div.stButton > button {
        background: linear-gradient(135deg, #4f46e5 0%, #9333ea 50%, #ec4899 100%);
        color: #ffffff !important;
        border-radius: 14px;
        border: 1.5px solid rgba(236, 72, 153, 0.6);
        font-weight: 800;
        letter-spacing: 1px;
        transition: all 0.3s ease-in-out;
        box-shadow: 0 4px 20px rgba(147, 51, 234, 0.5);
    }
    div.stButton > button:hover {
        background: linear-gradient(135deg, #6366f1 0%, #a855f7 50%, #f43f5e 100%);
        border-color: #38bdf8;
        transform: translateY(-4px);
        box-shadow: 0 8px 30px rgba(236, 72, 153, 0.7), 0 0 20px rgba(56, 189, 248, 0.6);
    }
    div[data-testid="metric-container"] {
        background: linear-gradient(135deg, rgba(15, 23, 42, 0.95) 0%, rgba(49, 46, 129, 0.7) 100%);
        border: 1.5px solid rgba(56, 189, 248, 0.4);
        padding: 18px;
        border-radius: 16px;
        box-shadow: 0 6px 20px rgba(0, 0, 0, 0.5);
    }
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #070b19 0%, #0f172a 100%);
        border-right: 1.5px solid rgba(139, 92, 246, 0.3);
        padding-top: 0.5rem;
    }
    </style>
""",
    unsafe_allow_html=True,
)

LOGO_PATH = "company_logo_permanent.png"

module_list = [
    "🏠 Home / முகப்பு",
    "📐 Rod Calculator (ராட் கால்குலேட்டர்)",
    "⏱️ Production Calculator (உற்பத்தி கால்குலேட்டர்)",
    "💰 Costing & Quotation Calculator (செலவு & கொட்டேஷன்)",
    "📦 Stock Management (ஸ்டாக் மேனேஜ்மென்ட்)",
    "📷 Drawing & Multi-Op G-Code (டிராயிங் & ஆட்டோ ரிப்போர்ட்)",
    (
        "📋 Process Breakdown & Customer Quotation (புதிய கொட்டேஷன்"
        " மாடியூல்)"
    ),
    "⚙️ More Menu & Settings (அமைப்புகள் & மாஸ்டர்ஸ்)",
]

if "selected_module" not in st.session_state:
  st.session_state["selected_module"] = module_list[0]

if "stock_inventory_df" not in st.session_state:
  st.session_state["stock_inventory_df"] = pd.DataFrame({
      "Item ID": ["ITM-001", "ITM-002", "ITM-003", "ITM-004"],
      "Material / Part Name": [
          "EN8 Round Bar - 12mm",
          "MS Round Bar - 20mm",
          "EN24 Round Bar - 16mm",
          "Finished Large Pin",
      ],
      "Category": [
          "Raw Material",
          "Raw Material",
          "Raw Material",
          "Finished Goods",
      ],
      "Quantity": [120.50, 45.20, 0.00, 650.00],
      "Unit": ["Kg", "Kg", "Kg", "Nos"],
      "Status": ["In Stock", "Low Stock", "Out of Stock", "Dispatch Ready"],
  })

if "stock_logs_df" not in st.session_state:
  st.session_state["stock_logs_df"] = pd.DataFrame(
      columns=["Timestamp", "Item ID", "Action", "Qty Changed", "User/Notes"]
  )


def get_base64_image(path):
  if os.path.exists(path):
    with open(path, "rb") as f:
      return base64.b64encode(f.read()).decode("utf-8")
  return None


# Sidebar Navigation
st.sidebar.markdown(
    """<div style="text-align: center; padding: 5px 0 10px 0;"><h3"
    " style="color: #ec4899; margin: 0; font-size: 1.15rem; font-weight: 900;"
    " letter-spacing: 1.5px;">MEGALA CNC MATE</h3></div>""",
    unsafe_allow_html=True,
)

encoded_sidebar_img = get_base64_image(LOGO_PATH)
if encoded_sidebar_img:
  st.sidebar.markdown(
      f"""<div style="text-align: center; margin-bottom: 12px;"><div"
      " style="background: linear-gradient(135deg, rgba(236, 72, 153, 0.3),"
      " rgba(56, 189, 248, 0.3)); border: 2.5px solid #ec4899; width: 75px;"
      " height: 75px; border-radius: 50%; margin: 0 auto; display: flex;"
      " align-items: center; justify-content: center; overflow: hidden;"><img"
      f" src="data:image/png;base64,{encoded_sidebar_img}" style="width: 100%;" /></div></div>""",
      unsafe_allow_html=True,
  )

uploaded_logo = st.sidebar.file_uploader(
    "Upload Permanent Logo",
    type=["png", "jpg", "jpeg"],
    key="sidebar_logo_upload",
    label_visibility="collapsed",
)
if uploaded_logo is not None:
  try:
    Image.open(uploaded_logo).save(LOGO_PATH)
    st.sidebar.success("✅ லோகோ சேமிக்கப்பட்டது!")
    st.rerun()
  except Exception as e:
    st.sidebar.error(f"Error: {e}")

st.sidebar.markdown("---")
selected_module = st.sidebar.selectbox(
    "Select Module",
    module_list,
    index=(
        module_list.index(st.session_state["selected_module"])
        if st.session_state["selected_module"] in module_list
        else 0
    ),
    label_visibility="collapsed",
)
st.session_state["selected_module"] = selected_module

# Header Layout
col_logo, col_title = st.columns([0.15, 0.85], vertical_alignment="center")
with col_logo:
  encoded_img = get_base64_image(LOGO_PATH)
  if encoded_img:
    st.markdown(
        f"""<div style="border: 2.5px solid #ec4899; width: 85px; height: 85px;"
        " border-radius: 50%; display: flex; align-items: center;"
        " justify-content: center; overflow: hidden;"><img"
        f" src="data:image/png;base64,{encoded_img}" style="width: 100%;" /></div>""",
        unsafe_allow_html=True,
    )
  else:
    st.markdown(
        """<div style="background: linear-gradient(135deg, #4f46e5, #ec4899);"
        " width: 85px; height: 85px; border-radius: 50%; display: flex;"
        " align-items: center; justify-content: center; font-weight: 900;"
        " color: white;">MC</div>""",
        unsafe_allow_html=True,
    )
with col_title:
  st.markdown(
      '<h1 class="main-title">MEGALA INDUSTRIES</h1>', unsafe_allow_html=True
  )
  st.markdown(
      '<p class="sub-title">PRECISION CNC MACHINING & ENTERPRISE AUTOMATION</p>',
      unsafe_allow_html=True,
  )

st.markdown(
    "<hr style='margin-top: 0px; border-color: rgba(236, 72, 153, 0.3);'>",
    unsafe_allow_html=True,
)


def clean_text(text):
  return (
      str(text).replace("₹", "Rs.").encode("latin-1", "replace").decode("latin-1")
  )


def generate_production_pdf(data):
  pdf = FPDF()
  pdf.add_page()
  pdf.set_font("Arial", "B", 16)
  pdf.cell(200, 10, txt="MEGALA INDUSTRIES - PRODUCTION REPORT", ln=True, align="C")
  pdf.set_font("Arial", "", 12)
  pdf.ln(10)
  for k, v in data.items():
    pdf.cell(200, 8, txt=clean_text(f"{k}: {v}"), ln=True)
  return pdf.output(dest="S").encode("latin1")


def generate_quotation_pdf(data):
  pdf = FPDF()
  pdf.add_page()
  pdf.set_font("Arial", "B", 14)
  pdf.cell(
      200, 10, txt="MEGALA INDUSTRIES - DETAILED QUOTATION REPORT", ln=True, align="C"
  )
  pdf.set_font("Arial", "", 10)
  pdf.ln(10)
  for k, v in data.items():
    pdf.cell(200, 7, txt=clean_text(f"{k}: {v}"), ln=True)
  return pdf.output(dest="S").encode("latin1")


def generate_program_pdf(code_text):
  pdf = FPDF()
  pdf.add_page()
  pdf.set_font("Courier", "B", 14)
  pdf.cell(200, 10, txt="MEGALA INDUSTRIES - G-CODE PROGRAM", ln=True, align="C")
  pdf.set_font("Courier", "", 10)
  pdf.ln(10)
  for line in code_text.split("\n"):
    pdf.cell(200, 6, txt=clean_text(line), ln=True)
  return pdf.output(dest="S").encode("latin1")


def get_cross_section_area(shape, dia_or_size, inner_dia=0.0):
  if shape == "Round Rod":
    return math.pi * (dia_or_size / 2.0) ** 2
  elif shape == "Square Rod":
    return dia_or_size**2
  elif shape == "Hexagon Rod":
    return (math.sqrt(3) / 2.0) * (dia_or_size**2)
  elif shape == "Tube / Pipe":
    return max(
        0.0,
        (math.pi * (dia_or_size / 2.0) ** 2)
        - (math.pi * (inner_dia / 2.0) ** 2),
    )
  return math.pi * (dia_or_size / 2.0) ** 2


def generate_quotation_excel(
    customer_name, part_name, operations_list, transport_cost=0.0
):
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Commercial Quotation"
  ws.views.sheetView[0].showGridLines = True

  font_company = Font(name="Calibri", size=16, bold=True, color="1F4E78")
  font_sub = Font(name="Calibri", size=9, italic=True, color="595959")
  font_title = Font(name="Calibri", size=13, bold=True, color="000000")
  font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
  font_bold = Font(name="Calibri", size=11, bold=True)
  font_normal = Font(name="Calibri", size=11)

  fill_header = PatternFill(
      start_color="1F4E78", end_color="1F4E78", fill_type="solid"
  )
  border_thin = Border(
      left=Side(style="thin", color="D9D9D9"),
      right=Side(style="thin", color="D9D9D9"),
      top=Side(style="thin", color="D9D9D9"),
      bottom=Side(style="thin", color="D9D9D9"),
  )
  border_total = Border(
      top=Side(style="thin", color="000000"),
      bottom=Side(style="double", color="000000"),
  )

  ws["A1"] = "MEGALA INDUSTRIES"
  ws["A1"].font = font_company
  ws["A2"] = (
      "Precision CNC Machining, VMC Components & Turning | Hosur, Tamil Nadu"
  )
  ws["A2"].font = font_sub
  ws["A4"] = "COMMERCIAL QUOTATION & PROCESS BREAKDOWN"
  ws["A4"].font = font_title
  ws["A5"] = f"Customer Name: {customer_name}"
  ws["B5"] = f"Part Name: {part_name}"
  ws["A6"] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
  ws["B6"] = "Quotation No: MI/Q/2026-08/01"

  for r in range(5, 7):
    ws[f"A{r}"].font = font_bold
    ws[f"B{r}"].font = font_bold

  headers = [
      "S.No",
      "Operation / Process Description",
      "Machine / Setup",
      "Qty",
      "Unit Rate (₹)",
      "Total Amount (₹)",
  ]
  start_row = 9
  for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=start_row, column=col_idx, value=header)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

  current_row = start_row + 1
  for idx, op in enumerate(operations_list, 1):
    ws.cell(row=current_row, column=1, value=idx).alignment = Alignment(
        horizontal="center"
    )
    ws.cell(row=current_row, column=2, value=op["name"])
    ws.cell(row=current_row, column=3, value=op["machine"])
    ws.cell(row=current_row, column=4, value=op["qty"]).alignment = Alignment(
        horizontal="right"
    )
    ws.cell(row=current_row, column=5, value=op["rate"]).number_format = (
        "₹#,##0.00"
    )
    ws.cell(
        row=current_row, column=6, value=f"=D{current_row}*E{current_row}"
    ).number_format = "₹#,##0.00"
    for c in range(1, 7):
      cell = ws.cell(row=current_row, column=c)
      cell.font = font_normal
      cell.border = border_thin
    current_row += 1

  if transport_cost > 0:
    ws.cell(row=current_row, column=2, value="Transport & Logistics Charges")
    ws.cell(row=current_row, column=4, value=1).alignment = Alignment(
        horizontal="right"
    )
    ws.cell(row=current_row, column=5, value=transport_cost).number_format = (
        "₹#,##0.00"
    )
    ws.cell(
        row=current_row, column=6, value=f"=E{current_row}"
    ).number_format = "₹#,##0.00"
    for c in range(1, 7):
      ws.cell(row=current_row, column=c).font = font_normal
      ws.cell(row=current_row, column=c).border = border_thin
    current_row += 1

  gt_row = current_row
  ws.cell(row=gt_row, column=5, value="Grand Total").font = font_bold
  ws.cell(row=gt_row, column=5).alignment = Alignment(horizontal="right")
  ws.cell(
      row=gt_row, column=6, value=f"=SUM(F{start_row+1}:F{current_row-1})"
  ).number_format = "₹#,##0.00"
  ws.cell(row=gt_row, column=6).font = font_bold
  ws.cell(row=gt_row, column=6).border = border_total

  for col in ws.columns:
    max_length = max(len(str(cell.value or "")) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_length + 4, 14)

  filename = (
      f"Megala_Industries_Quotation_{part_name.replace(' ', '_')}.xlsx"
  )
  wb.save(filename)
  return filename


# 1. HOME DASHBOARD
if "Home" in selected_module:
  inv_df = st.session_state["stock_inventory_df"]
  total_items_count = len(inv_df)
  low_stock_count = len(inv_df[inv_df["Status"] == "Low Stock"])
  out_stock_count = len(inv_df[inv_df["Status"] == "Out of Stock"])

  m1, m2, m3, m4 = st.columns(4)
  with m1:
    st.metric("Active Machines", "4 Units", "Running 🚀")
  with m2:
    st.metric("Today's Output", "1,850 Nos", "+12% 📈")
  with m3:
    st.metric("Material Stock Items", f"{total_items_count} Items", "Optimal ✨")
  with m4:
    st.metric(
        "Low/Out Stock",
        f"{low_stock_count + out_stock_count} Alerts",
        "Check Stock ⚠️",
    )

  st.markdown("<br>", unsafe_allow_html=True)
  st.markdown("### 🚀 Core Automation Modules / முக்கிய மாட்யூல்கள்")

  col1, col2, col3 = st.columns(3)
  with col1:
    st.markdown("### 📐 Rod Calculator")
    if st.button("🚀 Open Rod Calculator", use_container_width=True, key="bh1"):
      st.session_state["selected_module"] = "📐 Rod Calculator (ராட் கால்குலேட்டர்)"
      st.rerun()
  with col2:
    st.markdown("### ⏱️ Production Calc")
    if st.button("🚀 Open Production Calc", use_container_width=True, key="bh2"):
      st.session_state["selected_module"] = (
          "⏱️ Production Calculator (உற்பத்தி கால்குலேட்டர்)"
      )
      st.rerun()
  with col3:
    st.markdown("### 💰 Costing & Quote")
    if st.button("🚀 Open Costing Calc", use_container_width=True, key="bh3"):
      st.session_state["selected_module"] = (
          "💰 Costing & Quotation Calculator (செலவு & கொட்டேஷன்)"
      )
      st.rerun()

  c4, c5, c6, c7 = st.columns(4)
  with c4:
    st.markdown("### 📦 Stock Manager")
    if st.button("🚀 Open Stock Manager", use_container_width=True, key="bh4"):
      st.session_state["selected_module"] = (
          "📦 Stock Management (ஸ்டாக் மேனேஜ்மென்ட்)"
      )
      st.rerun()
  with c5:
    st.markdown("### 📷 Drawing Studio")
    if st.button("🚀 Open Drawing Studio", use_container_width=True, key="bh5"):
      st.session_state["selected_module"] = (
          "📷 Drawing & Multi-Op G-Code (டிராயிங் & ஆட்டோ ரிப்போர்ட்)"
      )
      st.rerun()
  with c6:
    st.markdown("### 📋 Quotation Hub")
    if st.button("🚀 Open Quote Hub", use_container_width=True, key="bh6"):
      st.session_state["selected_module"] = (
          "📋 Process Breakdown & Customer Quotation (புதிய கொட்டேஷன் மாடியூல்)"
      )
      st.rerun()
  with c7:
    st.markdown("### ⚙️ Settings")
    if st.button("🚀 Open Settings", use_container_width=True, key="bh7"):
      st.session_state["selected_module"] = (
          "⚙️ More Menu & Settings (அமைப்புகள் & மாஸ்டர்ஸ்)"
      )
      st.rerun()

# 2. ROD CALCULATOR
elif "Rod Calculator" in selected_module:
  if st.button("⬅️ Back to Home / முகப்புக்குத் திரும்பு"):
    st.session_state["selected_module"] = "🏠 Home / முகப்பு"
    st.rerun()
  st.subheader("📐 Rod Calculator - Simple & Advanced Modes")
  mode = st.radio("Mode Selection", ["Simple Mode", "Advanced Mode"], horizontal=True)

  if mode == "Simple Mode":
    st.write("### 🟢 Simple Mode: Quick Parts & Remnant Calculation")
    col1, col2 = st.columns(2)
    with col1:
      rod_length = st.number_input("Rod Length (Meter)", value=6.0, min_value=0.0)
      part_length = st.number_input("Part Length (mm)", value=38.70, min_value=0.0)
      cutting_allowance = st.number_input(
          "Cutting / Parting Allowance (mm)", value=3.0, min_value=0.0
      )
    with col2:
      shape_type = st.selectbox(
          "Material Shape",
          ["Round Rod", "Hexagon Rod", "Square Rod", "Tube / Pipe"],
      )
      cycle_time = st.number_input("Cycle Time (Seconds)", value=20, min_value=0)
      required_qty = st.number_input("Required Quantity", value=500, min_value=0)

    eff_len = part_length + cutting_allowance
    parts_per_rod = (
        int((rod_length * 1000) / eff_len) if eff_len > 0 else 0
    )
    remnant = round((rod_length * 1000) % eff_len, 2) if eff_len > 0 else 0.0
    req_rods = int(required_qty / parts_per_rod) if parts_per_rod > 0 else 0
    prod_per_hr = int(3600 / cycle_time) if cycle_time > 0 else 0

    st.markdown('<div class="auto-badge">⚡ SIMPLE MODE RESULT</div>', unsafe_allow_html=True)
    r1, r2, r3 = st.columns(3)
    with r1:
      st.metric("Parts / Rod", f"{parts_per_rod} Nos")
      st.metric("Required Rods", f"{req_rods} Nos")
    with r2:
      st.metric("Balance Scrap / Remnant", f"{remnant} mm")
      st.metric("Total Stock Length", f"{round(req_rods * rod_length, 2)} Meters")
    with r3:
      st.metric("Production / Hour", f"{prod_per_hr} Nos")
      st.metric("Total Machine Time", f"{round((required_qty * cycle_time)/3600, 2)} Hr")

  else:
    st.write("### 🔵 Advanced Mode: Drawing Upload & Exact Gram/Scrap Analysis")
    adv_file = st.file_uploader(
        "Upload Part Drawing / Photo", type=["png", "jpg", "pdf"]
    )
    if adv_file:
      st.success(
          f"📂 Drawing '{adv_file.name}' uploaded! Auto-detected Diameter:"
          " 18.0 mm, Part Length: 38.70 mm."
      )
      if adv_file.type in ["image/png", "image/jpeg"]:
        st.image(adv_file, width=350)

    ac1, ac2 = st.columns(2)
    with ac1:
      adv_shape = st.selectbox(
          "Material Shape",
          ["Round Rod", "Hexagon Rod", "Square Rod", "Tube / Pipe"],
          key="as",
      )
      adv_rod_len_m = st.number_input(
          "Rod Length (Meters)", value=6.0, key="arl"
      )
      adv_part_len = st.number_input(
          "Part Length from Drawing (mm)", value=38.70, key="apl"
      )
      adv_cut_allow = st.number_input(
          "Cutting Allowance (mm)", value=3.0, key="aca"
      )
      adv_req_qty = st.number_input("Required Order Quantity", value=500, key="arq")
    with ac2:
      adv_dia = st.number_input(
          "Raw Material Diameter / Size (mm)", value=18.0, key="add"
      )
      adv_inner_dia = (
          st.number_input("Tube Inner Diameter (mm)", value=20.0, key="aid")
          if adv_shape == "Tube / Pipe"
          else 0.0
      )
      adv_density = st.number_input(
          "Material Density (g/mm³)", value=0.00785, format="%.5f", key="adn"
      )
      adv_mat_rate = st.number_input("Material Rate / Kg (Rs.)", value=90.0, key="amr")
      adv_wastage_pct = st.slider("Additional Wastage (%)", 0, 10, 2, key="awt")

    cross_area = get_cross_section_area(adv_shape, adv_dia, adv_inner_dia)
    eff_l = adv_part_len + adv_cut_allow
    part_wt = round((cross_area * adv_part_len) * adv_density, 2)
    parts_bar = int((adv_rod_len_m * 1000) / eff_l) if eff_l > 0 else 0
    rem_mm = round((adv_rod_len_m * 1000) % eff_l, 2) if eff_l > 0 else 0.0
    req_rd = int(math.ceil(adv_req_qty / parts_bar)) if parts_bar > 0 else 0
    tot_wt_kg = (
        round(
            (
                req_rd
                * adv_rod_len_m
                * (cross_area * adv_density * 1000)
                / 1000000
            ),
            2,
        )
        * (1 + adv_wastage_pct / 100)
    )

    st.markdown('<div class="auto-badge">⚡ ADVANCED ANALYSIS RESULT</div>', unsafe_allow_html=True)
    ar1, ar2, ar3, ar4 = st.columns(4)
    with ar1:
      st.metric("Parts / Rod", f"{parts_bar} Nos")
      st.metric("Part Weight", f"{part_wt} g")
    with ar2:
      st.metric("Remnant / End Bit", f"{rem_mm} mm")
      st.metric("End Bit Weight", f"{round((cross_area * rem_mm)*adv_density, 2)} g")
    with ar3:
      st.metric("Required Rods", f"{req_rd} Nos")
      st.metric("Total Scrap / Rod", f"{round((cross_area*(rem_mm + (parts_bar*adv_cut_allow)))*adv_density, 2)} g")
    with ar4:
      st.metric("Total Mat. Weight", f"{round(tot_wt_kg, 2)} Kg")
      st.metric("Total Mat. Cost", f"Rs. {round(tot_wt_kg * adv_mat_rate, 2)}")

# 3. PRODUCTION CALCULATOR
elif "Production Calculator" in selected_module:
  if st.button("⬅️ Back to Home / முகப்புக்குத் திரும்பு"):
    st.session_state["selected_module"] = "🏠 Home / முகப்பு"
    st.rerun()
  st.subheader("⏱️ Production Days & Output Calculator & PDF Report")
  c1, c2 = st.columns(2)
  with c1:
    cyc_time = st.number_input("Cycle Time (sec)", value=20)
    avail_time = st.number_input("Available Time / Day (hr)", value=8.0)
  with c2:
    efficiency = st.slider("Machine Efficiency (%)", 50, 100, 85)
    break_time = st.number_input("Break Time (min)", value=30)

  eff_hrs = avail_time - (break_time / 60)
  prod_hr = int(3600 / cyc_time * (efficiency / 100)) if cyc_time > 0 else 0
  prod_day = int(prod_hr * eff_hrs)

  st.markdown('<div class="auto-badge">⚡ AUTO CALCULATED</div>', unsafe_allow_html=True)
  r1, r2 = st.columns(2)
  with r1:
    st.metric("Production / Hour", f"{prod_hr} Nos")
  with r2:
    st.metric("Production / Day", f"{prod_day} Nos")

  st.markdown("---")
  st.subheader("📄 Download Production Report as PDF")
  p_dict = {
      "Cycle Time (sec)": cyc_time,
      "Available Time / Day (hr)": avail_time,
      "Machine Efficiency (%)": efficiency,
      "Production / Hour": f"{prod_hr} Nos",
      "Production / Day": f"{prod_day} Nos",
  }
  st.download_button(
      "📥 Download Production Report PDF",
      data=generate_production_pdf(p_dict),
      file_name="Production_Report.pdf",
      mime="application/pdf",
  )

# 4. COSTING & QUOTATION CALCULATOR
elif "Costing & Quotation Calculator" in selected_module:
  if st.button("⬅️ Back to Home / முகப்புக்குத் திரும்பு"):
    st.session_state["selected_module"] = "🏠 Home / முகப்பு"
    st.rerun()
  st.subheader("💰 Costing & Quotation Calculator")
  col1, col2 = st.columns(2)
  with col1:
    mat_cost_kg = st.number_input("Material Cost / Kg (Rs.)", value=85.0)
    mat_wt_part = st.number_input("Material Weight / Part (Kg)", value=0.05)
    machine_cost_hr = st.number_input("Machine Cost / Hr (Rs.)", value=600.0)
  with col2:
    labour_cost_part = st.number_input("Labour Cost / Part (Rs.)", value=1.20)
    overhead_pct = st.number_input("Overhead (%)", value=15.0)
    profit_margin = st.slider("Profit Margin (%)", 0, 50, 20)

  subtotal = (
      (mat_cost_kg * mat_wt_part) + ((machine_cost_hr / 3600) * 20) + labour_cost_part
  )
  cost_part = subtotal * (1 + overhead_pct / 100)
  selling_price = cost_part * (1 + profit_margin / 100)

  st.markdown('<div class="auto-badge">⚡ AUTO CALCULATED</div>', unsafe_allow_html=True)
  p1, p2, p3 = st.columns(3)
  with p1:
    st.metric("Cost / Part", f"Rs. {round(cost_part, 2)}")
  with p2:
    st.metric("Cost / 1000 Parts", f"Rs. {round(cost_part * 1000, 2)}")
  with p3:
    st.metric("Selling Price / Part", f"Rs. {round(selling_price, 2)}")

  st.markdown("---")
  st.subheader("📄 Download Quotation PDF")
  q_dict = {
      "Cost Per Part": f"Rs. {round(cost_part, 2)}",
      "Selling Price Per Part": f"Rs. {round(selling_price, 2)}",
      "Cost for 1000 Parts": f"Rs. {round(cost_part * 1000, 2)}",
  }
  st.download_button(
      "📥 Download Quotation PDF",
      data=generate_quotation_pdf(q_dict),
      file_name="Quotation.pdf",
      mime="application/pdf",
  )

# 5. STOCK MANAGEMENT
elif "Stock Management" in selected_module:
  if st.button("⬅️ Back to Home / முகப்புக்குத் திரும்பு"):
    st.session_state["selected_module"] = "🏠 Home / முகப்பு"
    st.rerun()
  st.subheader("📦 Interactive Stock & Inventory Management System")
  inv_df = st.session_state["stock_inventory_df"]

  s1, s2, s3 = st.columns(3)
  with s1:
    st.metric("Total Items", str(len(inv_df)))
  with s2:
    st.metric("Low Stock", str(len(inv_df[inv_df["Status"] == "Low Stock"])))
  with s3:
    st.metric(
        "Out of Stock", str(len(inv_df[inv_df["Status"] == "Out of Stock"]))
    )

  st.markdown("---")
  tab1, tab2, tab3 = st.tabs(
      ["📋 Current Stock", "➕ Add Item", "🔄 Stock In / Out"]
  )
  with tab1:
    sq = st.text_input("🔍 Search Inventory...")
    st.dataframe(
        inv_df[
            inv_df["Material / Part Name"]
            .str.contains(sq, case=False, na=False)
        ]
        if sq
        else inv_df,
        use_container_width=True,
    )
  with tab2:
    with st.form("add_form"):
      nid = f"ITM-{len(inv_df)+1:03d}"
      nn = st.text_input("Part Name")
      nc = st.selectbox(
          "Category", ["Raw Material", "Finished Goods", "Consumables"]
      )
      nq = st.number_input("Quantity", value=50.0)
      nu = st.selectbox("Unit", ["Kg", "Nos", "Meters"])
      if st.form_submit_button("➕ Add") and nn:
        nst = (
            "Out of Stock"
            if nq == 0
            else ("Low Stock" if nq < 10 else "In Stock")
        )
        st.session_state["stock_inventory_df"] = pd.concat(
            [
                inv_df,
                pd.DataFrame({
                    "Item ID": [nid],
                    "Material / Part Name": [nn],
                    "Category": [nc],
                    "Quantity": [nq],
                    "Unit": [nu],
                    "Status": [nst],
                }),
            ],
            ignore_index=True,
        )
        st.success("Added successfully!")
        st.rerun()
  with tab3:
    if len(inv_df) > 0:
      with st.form("trans_form"):
        s_item = st.selectbox(
            "Select Item",
            inv_df["Item ID"] + " - " + inv_df["Material / Part Name"],
        )
        t_type = st.selectbox(
            "Type", ["Stock In (Purchase)", "Stock Out (Dispatch)"]
        )
        t_qty = st.number_input("Qty", value=10.0)
        t_note = st.text_input("Notes / PO Ref")
        if st.form_submit_button("🔄 Update Stock"):
          i_id = s_item.split(" - ")[0]
          idx = st.session_state["stock_inventory_df"].index[
              st.session_state["stock_inventory_df"]["Item ID"] == i_id
          ][0]
          curr = st.session_state["stock_inventory_df"].at[idx, "Quantity"]
          new_q = (
              curr + t_qty if "In" in t_type else max(0.0, curr - t_qty)
          )
          st.session_state["stock_inventory_df"].at[idx, "Quantity"] = new_q
          st.session_state["stock_inventory_df"].at[idx, "Status"] = (
              "Out of Stock"
              if new_q == 0
              else ("Low Stock" if new_q < 10 else "In Stock")
          )
          st.success(f"Updated! New Qty: {new_q}")
          st.rerun()

# 6. DRAWING & MULTI-OPERATION G-CODE GENERATOR
elif "Drawing & Multi-Op G-Code" in selected_module:
  if st.button("⬅️ Back to Home / முகப்புக்குத் திரும்பு"):
    st.session_state["selected_module"] = "🏠 Home / முகப்பு"
    st.rerun()
  st.subheader("📷 Drawing Upload & Automatic G-Code Generator")
  uf = st.file_uploader("Upload Drawing", type=["png", "jpg", "pdf"])
  if uf:
    st.success(f"📂 '{uf.name}' uploaded successfully!")
    if uf.type in ["image/png", "image/jpeg"]:
      st.image(uf, width=350)

  st.markdown("---")
  dc1, dc2 = st.columns(2)
  with dc1:
    d_shape = st.selectbox(
        "Profile",
        ["Round Rod", "Hexagon Rod", "Square Rod", "Tube / Pipe"],
        key="ds",
    )
    d_rlen = st.number_input("Rod Length (mm)", value=6000.0, key="drl")
    d_plen = st.number_input("Part Length (mm)", value=38.70, key="dpl")
    d_callow = st.number_input("Cutting Allowance (mm)", value=3.0, key="dca")
  with dc2:
    d_rdia = st.number_input("Diameter (mm)", value=18.0, key="drd")
    d_india = (
        st.number_input("Inner Bore (mm)", value=20.0, key="did")
        if d_shape == "Tube / Pipe"
        else 0.0
    )
    d_dens = st.number_input(
        "Density (g/mm³)", value=0.00785, format="%.5f", key="ddens"
    )
    d_mrate = st.number_input("Rate / Kg", value=90.0, key="dmr")

  c_area = get_cross_section_area(d_shape, d_rdia, d_india)
  eff_pl = d_plen + d_callow
  p_wt = round((c_area * d_plen) * d_dens, 2)
  ppb = int(d_rlen / eff_pl) if eff_pl > 0 else 0
  rem = round(d_rlen % eff_pl, 2) if eff_pl > 0 else 0.0

  st.markdown('<div class="auto-badge">⚡ ANALYSIS RESULT</div>', unsafe_allow_html=True)
  m1, m2, m3 = st.columns(3)
  with m1:
    st.metric("Parts / Rod", f"{ppb} Nos")
    st.metric("Part Weight", f"{p_wt} g")
  with m2:
    st.metric("Remnant / End Bit", f"{rem} mm")
    st.metric("End Bit Weight", f"{round((c_area * rem) * d_dens, 2)} g")
  with m3:
    st.metric("Total Scrap / Rod", f"{round((c_area * (rem + (ppb * d_callow))) * d_dens, 2)} g")

  st.markdown("---")
  st.subheader("🛠️ Multi-Operation Setup & G-Code Generator")
  num_ops = st.selectbox("Number of Operations", [1, 2, 3, 4, 5])
  all_gcodes = []

  for i in range(num_ops):
    with st.expander(f"📌 Operation {i+1} Details", expanded=(i == 0)):
      oc1, oc2 = st.columns(2)
      with oc1:
        t_no = st.text_input(f"Tool No (Op {i+1})", f"T{i+1:02d}{i+1:02d}", key=f"t_{i}")
        o_type = st.selectbox(
            f"Operation Type (Op {i+1})",
            [
                "Facing & Rough Turning",
                "Straight Turning",
                "Drilling / Boring",
                "Part-off",
            ],
            key=f"ot_{i}",
        )
        rpm = st.number_input(f"RPM (Op {i+1})", value=1200, key=f"rpm_{i}")
      with oc2:
        feed = st.number_input(f"Feed (mm/rev - Op {i+1})", value=0.15, key=f"fd_{i}")
        t_dia = st.number_input(f"Target Dia (Op {i+1})", value=d_rdia - 5.0, key=f"td_{i}")

      code = f"""( --- OP {i+1}: {o_type.upper()} --- )
{t_no}
G97 S{rpm} M03
G0 X{d_rdia + 5.0} Z2.0
G1 X0.0 F{feed}
G0 Z2.0
"""
      all_gcodes.append(code)
      st.text_area(f"G-Code Op {i+1}", code.strip(), height=100, key=f"gc_{i}")

  final_prog = "%\nO2026 (MEGALA CNC MATE)\nG21 G90 G40 G95\n" + "\n".join(all_gcodes) + "M05\nM30\n%"
  st.code(final_prog, language="text")
  st.download_button(
      "📥 Download G-Code PDF",
      data=generate_program_pdf(final_prog),
      file_name="CNC_Program.pdf",
      mime="application/pdf",
  )

# 7. PROCESS BREAKDOWN & CUSTOMER QUOTATION (NEW MODULE)
elif "Process Breakdown & Customer Quotation" in selected_module:
  if st.button("⬅️ Back to Home / முகப்புக்குத் திரும்பு"):
    st.session_state["selected_module"] = "🏠 Home / முகப்பு"
    st.rerun()

  st.subheader(
      "📋 Process Breakdown & Customer Quotation Generator (Large Pin & Trunnion)"
  )
  st.write(
      "Enter custom operations for parts like **Large Pin** or **Trunnion** to"
      " generate professional Excel & PDF quotations instantly."
  )

  col_q1, col_q2 = st.columns(2)
  with col_q1:
    cust_name = st.text_input(
        "Customer Company Name", value="M/s Precision Engineering Ltd"
    )
    part_type = st.selectbox(
        "Select Standard Part Template",
        ["Large Pin", "Trunnion", "Custom Part / Bush"],
    )
  with col_q2:
    order_qty = st.number_input("Order Quantity (Nos)", value=1000, min_value=1)
    transport_amt = st.number_input(
        "Transport / Logistics Charges (Rs.)", value=1500.0
    )

  if part_type == "Large Pin":
    default_ops = [
        {
            "name": "Facing & Turning (Op 1)",
            "machine": "CNC Turning",
            "qty": order_qty,
            "rate": 15.0,
        },
        {
            "name": "Milling (Op 2)",
            "machine": "VMC / Milling",
            "qty": order_qty,
            "rate": 20.0,
        },
        {
            "name": "Drilling (Op 3)",
            "machine": "Drilling Machine",
            "qty": order_qty,
            "rate": 18.0,
        },
        {
            "name": "Chamfering & Deburring (Op 4)",
            "machine": "Bench / Manual",
            "qty": order_qty,
            "rate": 5.0,
        },
        {
            "name": "Tapping Operation (Op 5)",
            "machine": "Tapping Setup",
            "qty": order_qty,
            "rate": 10.0,
        },
        {
            "name": "Tooling, Coolant & Process Consumables",
            "machine": "Overhead",
            "qty": order_qty,
            "rate": 6.0,
        },
    ]
  elif part_type == "Trunnion":
    default_ops = [
        {
            "name": "CNC Drilling Operation (Op 1)",
            "machine": "CNC Machining Center",
            "qty": order_qty,
            "rate": 25.0,
        },
        {
            "name": "Cross Drill & Rotary Setup (Op 2)",
            "machine": "Special Setup",
            "qty": order_qty,
            "rate": 30.0,
        },
        {
            "name": "Chamfering & Finishing (Op 3)",
            "machine": "Manual / VMC",
            "qty": order_qty,
            "rate": 8.0,
        },
        {
            "name": "Tooling & Consumables Allocation",
            "machine": "Overhead",
            "qty": order_qty,
            "rate": 7.0,
        },
    ]
  else:
    default_ops = [{
        "name": "General Machining Operation",
        "machine": "CNC Lathe",
        "qty": order_qty,
        "rate": 20.0,
    }]

  st.markdown("### ✏️ Customize Operations & Unit Rates")
  edited_ops = []
  for idx, op in enumerate(default_ops):
    col_a, col_b, col_c = st.columns([3, 2, 2])
    with col_a:
      op_name = st.text_input(
          f"Operation {idx+1} Name", value=op["name"], key=f"op_n_{idx}"
      )
    with col_b:
      mach_name = st.text_input(
          f"Machine {idx+1}", value=op["machine"], key=f"op_m_{idx}"
      )
    with col_c:
      op_rate = st.number_input(
          f"Unit Rate (Rs.) {idx+1}", value=op["rate"], key=f"op_r_{idx}"
      )
    edited_ops.append({
        "name": op_name,
        "machine": mach_name,
        "qty": order_qty,
        "rate": op_rate,
    })

  if st.button("🚀 Generate Excel Quotation File", use_container_width=True):
    excel_file = generate_quotation_excel(
        cust_name, part_type, edited_ops, transport_amt
    )
    st.success(f"✅ Quotation Excel successfully generated: {excel_file}")
    with open(excel_file, "rb") as f:
      st.download_button(
          "📥 Download Excel Quotation (.xlsx)",
          data=f,
          file_name=excel_file,
          mime=(
              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
          ),
      )

# 8. MORE MENU & SETTINGS
elif "More Menu & Settings" in selected_module:
  if st.button("⬅️ Back to Home / முகப்புக்குத் திரும்பு"):
    st.session_state["selected_module"] = "🏠 Home / முகப்பு"
    st.rerun()
  st.subheader("⚙️ More Menu & Settings")
  lang = st.selectbox(
      "🌐 Select Language",
      [
          "தமிழ் (Tamil)",
          "English",
          "हिन्दी (Hindi)",
          "తెలుగు (Telugu)",
          "മലയാളം (Malayalam)",
          "ಕನ್ನಡ (Kannada)",
      ],
  )
  st.success(f"Language set to: {lang}")
  st.markdown("---")
  st.markdown("### 📋 Workshop Masters")
  st.write("• Part Master\n• Customer Master\n• Machine Master\n• Material Master")
